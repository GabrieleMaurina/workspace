from openpyxl import load_workbook
import geocoder
from urllib.request import urlopen
from tkinter import Tk
from tkinter.ttk import Progressbar
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter.messagebox import showinfo, showerror
from functools import lru_cache
from os import getcwd
from json import loads
from datetime import timedelta

START_COL = 'K'
DEST_COL = 'F'
DIST_COL = 'N'
TIME_COL = 'O'
SHEET = 'MATRICE'

def try_n_times(f, n=3, default=None):
    for i in range(n):
        try:
            return f()
        except Exception:
            pass
    return default

@lru_cache(maxsize=None)
def get_coords(address):
    return tuple(geocoder.osm(address + ' italia').latlng[::-1])

@lru_cache(maxsize=None)
def get_dist_time(coords1, coords2):
    resource = urlopen(f'http://router.project-osrm.org/route/v1/car/{coords1[0]},{coords1[1]};{coords2[0]},{coords2[1]}?overview=full', timeout=2)
    route =  loads(resource.read().decode(resource.headers.get_content_charset()))['routes'][0]
    dist = str(round(route['distance']/1000, 1)).replace('.',',')
    time = str(timedelta(seconds=int(route['duration'])))[:-3]
    return dist, time

def get_files():
    Tk().withdraw()
    in_file = askopenfilename(title='Open Excel File', defaultextension='.xlsx', filetypes=(('Excel','.xlsx .xls'),), initialdir=getcwd())
    if in_file:
        out_file = asksaveasfilename(title='Save Excel File', defaultextension='.xlsx', filetypes=(('Excel','.xlsx .xls'),), initialdir=getcwd())
        if out_file:
            return in_file, out_file
    raise Exception('You must select an Excel file')

def main():
    try:    
        in_file, out_file = get_files()
        wb = load_workbook(filename = in_file)
        ws = wb[SHEET]

        start_col = map(lambda v: v.value, ws[START_COL][2:-1])
        dest_col = map(lambda v: v.value, ws[DEST_COL][2:-1])
        dist_col = ws[DIST_COL][2:-1]
        time_col = ws[TIME_COL][2:-1]

        done = 0
        tot = len(dist_col)
        
        for start, dest, dist, time in zip(start_col, dest_col, dist_col, time_col):
            if start and dest and not dist.value and not time.value:
                d, t = try_n_times(lambda : get_dist_time(get_coords(start), get_coords(dest)), default=(None, None))
                if d is not None and t is not None:
                    dist.value, time.value = d, t
                    done += 1
            
        wb.save(filename = out_file)
        showinfo(title='Completed', message=f'Computed {done}/{tot}, Excel file saved')
    except Exception as e:
        showerror(title='Error', message=str(e))

if __name__ == '__main__':
    main()
