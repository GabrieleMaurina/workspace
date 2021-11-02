from openpyxl import load_workbook
import googlemaps
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter.messagebox import showinfo

START_COL = 'K'
DEST_COL = 'F'
DIST_COL = 'N'
TIME_COL = 'O'
SHEET = 'MATRICE'
API_KEY = 'AIzaSyATpT-vZJyIpbFWCSe7Oc9RyKdTIv7AnjY'

Tk().withdraw()
in_file = askopenfilename(title='Open Excel File', defaultextension='.xlsx', filetypes=(('Excel','.xlsx .xls'),))
out_file = asksaveasfilename(title='Save Excel File', defaultextension='.xlsx', filetypes=(('Excel','.xlsx .xls'),))
showinfo(title='Completed', message='File saved')

gmaps = googlemaps.Client(key=API_KEY)

wb = load_workbook(filename = in_file)
ws = wb.get_sheet_by_name(SHEET)

start_col = map(lambda v: v.value, ws[START_COL][1:])
dest_col = map(lambda v: v.value, ws[DEST_COL][1:])
dist_col = ws[DIST_COL][1:]
time_col = ws[TIME_COL][1:]

for start, dest, dist, time in zip(start_col, dest_col, dist_col, time_col):
    if start and dest and not dist.value and not time.value:
        res = gmaps.distance_matrix(start, dest)
        print(res)
        exit(0)
        dist.value = '#'

wb.save(filename = out_file)
